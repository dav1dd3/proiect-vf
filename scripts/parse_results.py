from __future__ import annotations
import re
import csv
from pathlib import Path
from typing import Any, Dict, List, Optional

CPU_RE  = re.compile(r"CPU time\s*:\s*([\d.]+)\s*s", re.IGNORECASE)
MEM_RE  = re.compile(r"Memory used\s*:\s*([\d.]+)\s*MB", re.IGNORECASE)
CONF_RE = re.compile(r"^\s*conflicts\s*:\s*(\d+)", re.IGNORECASE | re.MULTILINE)
DEC_RE  = re.compile(r"^\s*decisions\s*:\s*(\d+)", re.IGNORECASE | re.MULTILINE)
PROP_RE = re.compile(r"^\s*propagations\s*:\s*(\d+)", re.IGNORECASE | re.MULTILINE)

FIELDS = [
    "instance", "os", "status",
    "cpu_time_s", "memory_mb",
    "conflicts", "decisions", "propagations",
    "conflicts_per_sec", "propagations_per_sec",
]

def detect_os(stem_lower: str) -> str:
    if "windows" in stem_lower or "win" in stem_lower:
        return "windows"
    if "macos" in stem_lower or "osx" in stem_lower or "mac" in stem_lower:
        return "macos"
    if "linux" in stem_lower:
        return "linux"
    return "unknown"

def strip_os_suffix(stem: str) -> str:
    s = stem
    for suf in ["_windows", "_win", "_macos", "_osx", "_mac", "_linux"]:
        if s.lower().endswith(suf):
            s = s[: -len(suf)]
    return s

def find1(regex: re.Pattern, text: str) -> str:
    m = regex.search(text)
    return m.group(1) if m else ""

def to_number(s: Any) -> Optional[float | int]:
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return s
    s = str(s).strip()
    if s == "":
        return None
    try:
        if "." in s:
            return float(s)
        return int(s)
    except ValueError:
        return None

def parse_status(text: str) -> str:
    if "UNSATISFIABLE" in text:
        return "UNSATISFIABLE"
    if "SATISFIABLE" in text:
        return "SATISFIABLE"
    return "UNKNOWN"

def parse_log(path: Path) -> Dict[str, Any]:
    txt = path.read_text(errors="ignore")
    row: Dict[str, Any] = {}

    row["status"] = parse_status(txt)
    row["cpu_time_s"] = to_number(find1(CPU_RE, txt))
    row["memory_mb"] = to_number(find1(MEM_RE, txt))
    row["conflicts"] = to_number(find1(CONF_RE, txt))
    row["decisions"] = to_number(find1(DEC_RE, txt))
    row["propagations"] = to_number(find1(PROP_RE, txt))

    cpu = row.get("cpu_time_s")
    conf = row.get("conflicts")
    prop = row.get("propagations")

    row["conflicts_per_sec"] = (conf / cpu) if (cpu and conf) else None
    row["propagations_per_sec"] = (prop / cpu) if (cpu and prop) else None

    return row

def write_csv(rows: List[Dict[str, Any]], out_csv: Path) -> None:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS, delimiter=";")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in FIELDS})

def try_write_xlsx(rows: List[Dict[str, Any]], out_xlsx: Path) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        return False

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    ws.append(FIELDS)
    header_font = Font(bold=True)
    for col_idx, name in enumerate(FIELDS, start=1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([r.get(k, None) for k in FIELDS])

    ws.freeze_panes = "A2"

    for rr in ws.iter_rows(min_row=2):
        if rr[3].value is not None:
            rr[3].number_format = "0.00"
        if rr[4].value is not None:
            rr[4].number_format = "0.00"
        if rr[8].value is not None:
            rr[8].number_format = "0.00"
        if rr[9].value is not None:
            rr[9].number_format = "0.00"

    for col_idx, col_name in enumerate(FIELDS, start=1):
        max_len = len(col_name)
        for (val,) in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        # cap width
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 2, 60)

    wb.save(out_xlsx)
    return True

def plot_grouped_bars(rows: List[Dict[str, Any]], out_dir: Path) -> None:
    import matplotlib.pyplot as plt

    out_dir.mkdir(parents=True, exist_ok=True)

    data: Dict[str, Dict[str, Dict[str, Any]]] = {}
    for r in rows:
        inst = r.get("instance", "unknown")
        osn = r.get("os", "unknown")
        data.setdefault(inst, {})[osn] = r

    instances = sorted(data.keys())
    preferred_order = ["macos", "windows", "linux", "unknown"]
    os_list = sorted({osn for inst in data for osn in data[inst].keys()},
                     key=lambda x: preferred_order.index(x) if x in preferred_order else 999)

    def get_metric(inst: str, osn: str, metric: str) -> Optional[float]:
        v = data.get(inst, {}).get(osn, {}).get(metric, None)
        if v is None:
            return None
        try:
            return float(v)
        except Exception:
            return None

    def make_chart(metric: str, ylabel: str, filename: str) -> None:
        x = list(range(len(instances)))
        bar_width = 0.8 / max(1, len(os_list))

        plt.figure()
        for i, osn in enumerate(os_list):
            vals = [get_metric(inst, osn, metric) for inst in instances]
            plot_vals = [v if v is not None else 0.0 for v in vals]
            xs = [xi - 0.4 + (i + 0.5) * bar_width for xi in x]
            plt.bar(xs, plot_vals, width=bar_width, label=osn)

        plt.xticks(x, instances, rotation=15, ha="right")
        plt.ylabel(ylabel)
        plt.legend()
        plt.tight_layout()
        plt.savefig(out_dir / filename, dpi=200)
        plt.close()

    make_chart("cpu_time_s", "CPU time (s)", "cpu_time_bar.png")
    make_chart("conflicts_per_sec", "Conflicts / sec", "conflicts_per_sec_bar.png")
    make_chart("propagations_per_sec", "Propagations / sec", "propagations_per_sec_bar.png")

def main() -> None:
    project_root = Path(".")
    raw_dir = project_root / "results" / "raw"
    if not raw_dir.exists():
        raise SystemExit("Missing folder: results/raw\nPut your *.txt logs in results/raw/")

    rows: List[Dict[str, Any]] = []
    for p in sorted(raw_dir.glob("*.txt")):
        stem = p.stem  
        stem_lower = stem.lower()

        os_name = detect_os(stem_lower)

        base = strip_os_suffix(stem)
        instance = base + ".cnf" 
        if instance.lower().endswith(".cnf.cnf"):
            instance = instance[:-4]

        parsed = parse_log(p)
        row = {
            "instance": instance,
            "os": os_name,
            **parsed,
        }
        rows.append(row)

    rows.sort(key=lambda r: (str(r.get("instance","")), str(r.get("os",""))))

    out_csv = project_root / "results" / "results.csv"
    out_xlsx = project_root / "results" / "results.xlsx"

    write_csv(rows, out_csv)
    wrote_xlsx = try_write_xlsx(rows, out_xlsx)

    plot_grouped_bars(rows, project_root / "results")

    print(f"Generated CSV:  {out_csv}")
    if wrote_xlsx:
        print(f"Generated XLSX: {out_xlsx}")
    else:
        print("openpyxl not installed -> XLSX not generated (CSV + PNG charts generated).")
        print("Install with: python3 -m pip install openpyxl")

    print("Generated charts:")
    print("  results/cpu_time_bar.png")
    print("  results/conflicts_per_sec_bar.png")
    print("  results/propagations_per_sec_bar.png")

if __name__ == "__main__":
    main()